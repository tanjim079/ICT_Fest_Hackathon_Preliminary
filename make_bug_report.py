"""Generate a formatted Excel bug report for the CoWork API hackathon."""
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────
C_TITLE_BG   = "1F3864"   # dark navy
C_TITLE_FG   = "FFFFFF"
C_H1_BG      = "2E75B6"   # medium blue  (section headers)
C_H2_BG      = "BDD7EE"   # light blue   (column headers)
C_H2_FG      = "1F3864"
C_FIXED_BG   = "E2EFDA"   # light green  (fixed = Yes)
C_BUG_BG     = "FCE4D6"   # light orange (bug rows)
C_OK_BG      = "EBF3FB"   # pale blue    (no-bug rows)
C_BORDER     = "9DC3E6"

thin = Side(border_style="thin", color=C_BORDER)
full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(ws, row, col,
               value="", bold=False, italic=False, size=10,
               fg=None, bg=None, wrap=True, align_h="left", align_v="center",
               border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, italic=italic, size=size,
                  color=fg or "000000")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align_h, vertical=align_v,
                            wrap_text=wrap)
    if border:
        c.border = full_border
    return c

# ══════════════════════════════════════════════════════════════
# SHEET 1 — Summary
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"

ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 42
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 32
ws1.column_dimensions["E"].width = 18
ws1.column_dimensions["F"].width = 10

# Title row
ws1.merge_cells("A1:F1")
t = ws1.cell(row=1, column=1, value="CoWork API — Bug Report  |  ICT Fest Hackathon 2026")
t.font = Font(name="Calibri", bold=True, size=16, color=C_TITLE_FG)
t.fill = PatternFill("solid", fgColor=C_TITLE_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

# Sub-title
ws1.merge_cells("A2:F2")
s = ws1.cell(row=2, column=1, value="Total bugs found & fixed: 21   |   Smoke tests: PASSED ✅")
s.font = Font(name="Calibri", bold=True, size=11, color=C_H2_FG)
s.fill = PatternFill("solid", fgColor=C_H2_BG)
s.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 22

# Column headers
ws1.row_dimensions[3].height = 20
headers = ["#", "Tested Area / Bug Description", "Bug ID", "File", "Status", "Fixed"]
hbgs    = [C_H2_BG]*6
for ci, (h, bg) in enumerate(zip(headers, hbgs), 1):
    cell_style(ws1, 3, ci, h, bold=True, fg=C_H2_FG, bg=bg, align_h="center")

# Data rows
rows = [
    # (area, bug_id, file, status, fixed)
    ("Booking — start_time grace window (5-min past allowed)", "BUG-01", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — minimum 1-hour duration not enforced",          "BUG-02", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — overlap predicate used <= instead of <",        "BUG-03", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — TOCTOU race: conflict + quota not atomic",      "BUG-04", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — list sorted descending instead of ascending",   "BUG-05", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — pagination offset wrong + limit hardcoded 10",  "BUG-06", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — detail endpoint overwrites start_time with created_at", "BUG-07", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — member can view another member's booking (→ 404)", "BUG-08", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Refund — 0% refund tier never applied (else gave 50%)",          "BUG-09", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Refund — response amount used round() vs half-up ledger", "BUG-10", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("JWT — /auth/refresh rotation did not invalidate presented refresh JTI", "BUG-11", "routers/auth.py", "🐛 Bug", "✅ Fixed"),
    ("Stats — revenue goes negative after server restart",      "BUG-12", "services/stats.py",   "🐛 Bug", "✅ Fixed"),
    ("Booking — concurrent cancellation race (duplicate refund logs)", "BUG-13", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Register — concurrent registration race (500 database error)", "BUG-14", "routers/auth.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — availability cache is not invalidated when a booking is cancelled", "BUG-15", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — usage report cache is not invalidated when a booking is created", "BUG-16", "routers/bookings.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — usage report cache is not invalidated when a room is created", "BUG-17", "routers/rooms.py", "🐛 Bug", "✅ Fixed"),
    ("Booking — notification lock acquisition deadlock",        "BUG-18", "services/notifications.py", "🐛 Bug", "✅ Fixed"),
    ("DateTime — timezone-aware UTC offset stripped instead of converted", "BUG-19", "timeutils.py", "🐛 Bug", "✅ Fixed"),
    ("Export — include_all scoping bypass exposes other org bookings", "BUG-20", "services/export.py", "🐛 Bug", "✅ Fixed"),
    ("Rate Limiting — concurrent requests bypass rolling 60-second limit", "BUG-21", "services/ratelimit.py", "🐛 Bug", "✅ Fixed"),
    ("Login — credential validation logic",                     "—",      "routers/auth.py",     "✅ OK",  "N/A"),
    ("Refund Formula — half-up rounding",                       "—",      "services/refunds.py", "✅ OK",  "N/A"),
    ("Admin Report — includes rooms with zero bookings",        "—",      "routers/admin.py",    "✅ OK",  "N/A"),
]

for ri, (desc, bug_id, file_, status, fixed) in enumerate(rows, 4):
    is_bug = bug_id != "—"
    row_bg = C_BUG_BG if is_bug else C_OK_BG
    ws1.row_dimensions[ri].height = 18
    cell_style(ws1, ri, 1, ri - 3,    bold=True,  bg=row_bg, align_h="center")
    cell_style(ws1, ri, 2, desc,                  bg=row_bg)
    cell_style(ws1, ri, 3, bug_id,    bold=is_bug,bg=row_bg, align_h="center")
    cell_style(ws1, ri, 4, file_,     italic=True, bg=row_bg)
    status_fg = "C00000" if is_bug else "375623"
    cell_style(ws1, ri, 5, status,    bold=True, fg=status_fg, bg=row_bg, align_h="center")
    fix_bg = C_FIXED_BG if fixed == "✅ Fixed" else row_bg
    cell_style(ws1, ri, 6, fixed,     bold=True, bg=fix_bg, align_h="center")

# Freeze panes
ws1.freeze_panes = "A4"

# ══════════════════════════════════════════════════════════════
# SHEET 2 — Detailed Findings
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Detailed Findings")

ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 26
ws2.column_dimensions["D"].width = 22
ws2.column_dimensions["E"].width = 50
ws2.column_dimensions["F"].width = 50
ws2.column_dimensions["G"].width = 12

ws2.merge_cells("A1:G1")
t2 = ws2.cell(row=1, column=1, value="CoWork API — Detailed Bug Findings")
t2.font = Font(name="Calibri", bold=True, size=15, color=C_TITLE_FG)
t2.fill = PatternFill("solid", fgColor=C_TITLE_BG)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 34

col_hdrs = ["Bug ID", "Category", "File", "Location", "Description of Bug", "Fix Applied", "Status"]
ws2.row_dimensions[2].height = 20
for ci, h in enumerate(col_hdrs, 1):
    cell_style(ws2, 2, ci, h, bold=True, fg=C_H2_FG, bg=C_H2_BG, align_h="center")

details = [
    ("BUG-01","Booking","routers/bookings.py","L86 · create_booking()",
     "start_time grace window: `start <= now - timedelta(seconds=300)` accepted bookings up to 5 minutes in the past. Rule requires strictly in the future.",
     "Changed condition to `if start <= now`","✅ Fixed"),

    ("BUG-02","Booking","routers/bookings.py","L93 · create_booking()",
     "Missing minimum-duration check: only > MAX_DURATION_HOURS was enforced. A 0-hour booking was silently accepted. Duration must be 1–8 hours integer.",
     "Changed to `if duration_hours < MIN_DURATION_HOURS or duration_hours > MAX_DURATION_HOURS`","✅ Fixed"),

    ("BUG-03","Booking","routers/bookings.py","L50 · _has_conflict()",
     "Non-strict overlap predicate used <= on both sides. Back-to-back bookings (10:00-11:00 then 11:00-12:00) were incorrectly flagged as conflicts.",
     "Changed to `b.start_time < end and start < b.end_time` (strict inequalities)","✅ Fixed"),

    ("BUG-04","Booking / Concurrency","routers/bookings.py","L100-117 · create_booking()",
     "TOCTOU race: conflict check + quota check occurred outside a DB write-lock. Two simultaneous requests could both pass and both commit, causing double-booking or quota violations.",
     "Added `db.execute(text('BEGIN IMMEDIATE'))` before check+insert to acquire SQLite write-lock atomically","✅ Fixed"),

    ("BUG-05","Booking / List","routers/bookings.py","L137 · list_bookings()",
     "List sort order was descending (newest first). Spec requires ascending by start_time, then ascending by id.",
     "Changed to `Booking.start_time.asc(), Booking.id.asc()`","✅ Fixed"),

    ("BUG-06","Booking / Pagination","routers/bookings.py","L138-139 · list_bookings()",
     "Offset used `page * limit` (page 1 → skip 10 items, missing first page). Limit was hardcoded to 10 ignoring the `limit` query parameter.",
     "Changed to `.offset((page - 1) * limit).limit(limit)`","✅ Fixed"),

    ("BUG-07","Booking / Detail","routers/bookings.py","L166 · get_booking()",
     "`response['start_time'] = iso_utc(booking.created_at)` overwrote the correct start_time field with created_at in the detail response.",
     "Removed the erroneous overwrite line; serialize_booking() already sets start_time correctly","✅ Fixed"),

    ("BUG-08","Booking / Visibility","routers/bookings.py","L160-175 · get_booking()",
     "GET /bookings/{id} returned any booking in the org regardless of who owns it. Members must get 404 for another member's booking.",
     "Added ownership check: `if user.role != 'admin' and booking.user_id != user.id: raise AppError(404, ...)`","✅ Fixed"),

    ("BUG-09","Refund","routers/bookings.py","L201-206 · cancel_booking()",
     "0% refund tier never applied: the else branch assigned refund_percent = 50 instead of 0. Both elif and else returned 50%. Short-notice cancellations always refunded 50%.",
     "Fixed to three correct tiers: >=48h→100%, >=24h→50%, <24h→0% using float notice_hours","✅ Fixed"),

    ("BUG-10","Refund","routers/bookings.py","L208 · cancel_booking()",
     "Response used Python round() (banker's rounding) while log_refund() uses half-up formula (price*pct+50)//100. Values could differ for odd-cent amounts.",
     "Response now reads amount from log_refund() return value: `refund_entry = log_refund(...); refund_amount_cents = refund_entry.amount_cents`","✅ Fixed"),

    ("BUG-11","JWT / Auth","routers/auth.py","L87 · refresh()",
     "The /auth/refresh endpoint failed to revoke the presented refresh token, allowing a single refresh token to be reused multiple times.",
     "Restored the call to `revoke_access_token(data)` for the presented refresh token payload to invalidate it.","✅ Fixed"),

    ("BUG-12","Stats","services/stats.py","L30 · record_cancel()",
     "revenue - price_cents went negative after server restart (empty in-memory state). count was guarded with max(0,...) but revenue was not.",
     "Changed to `max(0, revenue - price_cents)`","✅ Fixed"),

    ("BUG-13","Booking / Cancellation Concurrency","routers/bookings.py","L192 · cancel_booking()",
     "Concurrent cancellation requests could race the status check before writing to the database, resulting in duplicate refund logs for a single cancellation.",
     "Added `db.execute(text('BEGIN IMMEDIATE'))` to acquire write lock before cancel transactions.","✅ Fixed"),

    ("BUG-14","Register / Concurrency","routers/auth.py","L23 · register()",
     "Concurrent user registration requests with identical details could bypass check and trigger a 500 error due to unique constraints.",
     "Added `db.execute(text('BEGIN IMMEDIATE'))` to serialize registration transactions.","✅ Fixed"),

    ("BUG-15","Booking / Cache Invalidation","routers/bookings.py","L236 · cancel_booking()",
     "When a booking is cancelled, the room availability cache was not invalidated, causing subsequent availability lookups to show stale busy intervals.",
     "Added cache.invalidate_availability(booking.room_id, booking.start_time.date().isoformat()) to cancel_booking.","✅ Fixed"),

    ("BUG-16","Booking / Cache Invalidation","routers/bookings.py","L129 · create_booking()",
     "When a booking is created, the usage report cache for the organization was not invalidated, keeping new bookings off the report until the cache aged out.",
     "Added cache.invalidate_report(user.org_id) to create_booking.","✅ Fixed"),

    ("BUG-17","Room / Cache Invalidation","routers/rooms.py","L56 · create_room()",
     "When a room is created, the usage report cache for the organization was not invalidated, keeping the new room off the report until the cache aged out.",
     "Added cache.invalidate_report(admin.org_id) to create_room.","✅ Fixed"),

    ("BUG-18","Booking / Concurrency","services/notifications.py","L24-36 · notify_created() / notify_cancelled()",
     "Simulated email and audit log locks were acquired in opposite orders in the creation and cancellation paths. Under concurrent load, this could trigger circular-wait deadlocks, freezing the server process.",
     "Unified the locking order in both routines to always acquire _email_lock before _audit_lock.","✅ Fixed"),

    ("BUG-19","DateTime","timeutils.py","L13 · parse_input_datetime()",
     "For timezone-aware datetimes, the offset was stripped (replaced with None) instead of correctly converting the datetime to UTC. This caused incorrect dates and times to be stored for non-UTC input.",
     "Changed the offset check to use astimezone(timezone.utc) before removing tzinfo.","✅ Fixed"),

    ("BUG-20","Export / Security","services/export.py","L44-53 · generate_export()",
     "Specifying include_all=true with a room_id bypassed the organization scoping checks, leaking booking details from other organizations.",
     "Enforced room ownership validation checking Room.org_id == org_id before executing raw room queries.","✅ Fixed"),

    ("BUG-21","Rate Limiting / Concurrency","services/ratelimit.py","L20 · record_check()",
     "Rate limit counters were kept in local process-level dictionaries, allowing concurrent requests across different worker processes to completely bypass the rolling 60-second limit.",
     "Migrated the rate-limiting storage to a database table ('rate_limit_logs') and serialized reads/writes under a SQLite IMMEDIATE transaction write lock.","✅ Fixed"),
]

for ri, row in enumerate(details, 3):
    bug_id, cat, file_, loc, desc, fix, status = row
    ws2.row_dimensions[ri].height = 56
    cell_style(ws2, ri, 1, bug_id,  bold=True, bg=C_BUG_BG, align_h="center")
    cell_style(ws2, ri, 2, cat,     bold=True, fg=C_H1_BG,  bg="EBF3FB")
    cell_style(ws2, ri, 3, file_,   italic=True, bg="F2F2F2")
    cell_style(ws2, ri, 4, loc,     bg="F2F2F2", align_h="center")
    cell_style(ws2, ri, 5, desc,    bg="FFF2CC")
    cell_style(ws2, ri, 6, fix,     bg=C_FIXED_BG)
    cell_style(ws2, ri, 7, status,  bold=True, fg="375623", bg=C_FIXED_BG, align_h="center")

ws2.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════
# SHEET 3 — Files Changed
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Files Changed")
ws3.column_dimensions["A"].width = 34
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 60

ws3.merge_cells("A1:C1")
t3 = ws3.cell(row=1, column=1, value="Files Modified")
t3.font = Font(name="Calibri", bold=True, size=14, color=C_TITLE_FG)
t3.fill = PatternFill("solid", fgColor=C_TITLE_BG)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

for ci, h in enumerate(["File", "Bugs Fixed", "Bug IDs"], 1):
    cell_style(ws3, 2, ci, h, bold=True, fg=C_H2_FG, bg=C_H2_BG, align_h="center")
ws3.row_dimensions[2].height = 18

file_data = [
    ("app/routers/bookings.py", 14, "BUG-01, BUG-02, BUG-03, BUG-04, BUG-05, BUG-06, BUG-07, BUG-08, BUG-09, BUG-10, BUG-13, BUG-15, BUG-16"),
    ("app/routers/rooms.py",     2, "BUG-17"),
    ("app/routers/auth.py",      2, "BUG-11, BUG-14"),
    ("app/services/stats.py",    1, "BUG-12"),
    ("app/services/notifications.py", 1, "BUG-18"),
    ("app/timeutils.py",         1, "BUG-19"),
    ("app/services/export.py",   1, "BUG-20"),
    ("app/services/ratelimit.py", 1, "BUG-21"),
]
for ri, (f, n, ids) in enumerate(file_data, 3):
    ws3.row_dimensions[ri].height = 20
    cell_style(ws3, ri, 1, f,   italic=True, bg="F2F2F2")
    cell_style(ws3, ri, 2, n,   bold=True, bg=C_BUG_BG, align_h="center")
    cell_style(ws3, ri, 3, ids, bg=C_FIXED_BG)

out = r"d:\project\ICT_Fest_Hackathon_Preliminary\bug_report.xlsx"
wb.save(out)
print(f"Saved: {out}")
